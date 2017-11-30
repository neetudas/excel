from openpyxl import load_workbook

#excel_load = load_workbook(filename = 'accelerate.xlsx')
wb = load_workbook(filename = 'accelerate.xlsx')
print ('All availabel sheets in workbook are:  ', wb.sheetnames)
sheet_opened = wb['Sheet']
year_count = 0
month_count = 0
week_count = 0
prod_year_count = 0
prod_month_count = 0
prod_week_count = 0
# print (sheet_opened.max_column)
for row in range(1, sheet_opened.max_row+1):
  if str(sheet_opened['A'+str(row)].value).lower() == 'year':
      year_count += sheet_opened['F'+str(row)].value 
      time_A = sheet_opened['B' + str(row)].value or 0
      time_C = sheet_opened['D' + str(row)].value or 0
      prod_year_count += time_A + time_C
      #print (year_count)  

  if str(sheet_opened['A'+str(row)].value).lower() == 'month':
      month_count += sheet_opened['F'+str(row)].value
      time_A = sheet_opened['B' + str(row)].value or 0
      time_C = sheet_opened['D' + str(row)].value or 0
      prod_month_count += time_A + time_C

  if str(sheet_opened['A'+str(row)].value).lower() == 'week':
      week_count += sheet_opened['F'+str(row)].value
      time_A = sheet_opened['B' + str(row)].value or 0
      time_C = sheet_opened['D' + str(row)].value or 0
      prod_week_count += time_A +time_C

for row in range(1, sheet_opened.max_row+1):  
  if str(sheet_opened['A'+str(row)].value).lower() == 'year count':
      sheet_opened['B'+str(row)] = year_count
  if str(sheet_opened['A'+str(row)].value).lower() == 'month count':
      sheet_opened['B'+str(row)] = month_count
  if str(sheet_opened['A'+str(row)].value).lower() == 'week count':
      sheet_opened['B'+str(row)] = week_count
  if str(sheet_opened['A'+str(row)].value).lower() == 'prod-year':
      sheet_opened['B'+str(row)] = year_count / prod_year_count
  if str(sheet_opened['A'+str(row)].value).lower() == 'prod-month':
      sheet_opened['B'+str(row)] = year_count / prod_month_count
  if str(sheet_opened['A'+str(row)].value).lower() == 'prod-week':
      sheet_opened['B'+str(row)] = year_count / prod_week_count

wb.save(filename = 'accelerate.xlsx')
      


